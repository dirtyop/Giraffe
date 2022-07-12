import xlsxwriter
import xlrd
import openpyxl
from openpyxl import workbook
"""
HOST = "localhost"
USER = "schnell"
DBNAME = "chhattisgarh"
PASSWORD = "6Pn>/<U!0^k,:+"
conn = MySQLdb.connect(HOST, USER, PASSWORD, DBNAME)
cur = conn.cursor()
query = "select * from energy where " \
       "status = 1 and project_type ='eesl' order by project_name"
cur.execute(query)
prolist = cur.fetchall()
file1 = open("nwfile.csv", 'a')
"""
prolist = [12, 23, 15, 16, 14, 20, 23, 24, 26, 28]
workbook = xlsxwriter.Workbook('xlfile1.xlsx')
worksheet = workbook.add_worksheet()
row = 0
coulumn = 0
for i in prolist:
    #print(i)
    if i != "":
        #print(type(i))
        worksheet.write(row, coulumn, i)
        coulumn += 1

workbook.close()
location = "C:\\Users\\anand\Documents\EnergySaving\Monthly_Savings_Report_Jan_2020.xlsx"
file1 = openpyxl.load_workbook(location)
#if file1.sheetnames == "SAVINGS REPORT":
for sheet in file1:
    print(sheet.title)
    if sheet.title == "Energy Report":
        for rows in sheet.values:
            print(rows)
            #for value in rows:
                #print(value)
"""       
file1 = open("nwfile.csv", 'a')
file1.write(str(prolist) + '\n')
file1.close()
print(prolist)

"""

